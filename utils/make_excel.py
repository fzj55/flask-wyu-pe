import datetime

import openpyxl as vb
from openpyxl import styles

side = vb.styles.Side(style='thin', color='FF000000')
border = vb.styles.Border(left=side, right=side, top=side, bottom=side)
title_font = styles.Font(name=u'微软雅黑', bold=True, size=20)
subheading_font = styles.Font(name=u'微软雅黑', size=16)
normal_font=styles.Font(name=u'宋体', size=16)
align = vb.styles.alignment.Alignment(horizontal='center', vertical='center')


def make_timetable(data:dict):
    excel = vb.Workbook()
    table = excel['Sheet']

    table.column_dimensions['A'].width = 39.22
    table.column_dimensions['B'].width = 39.22

    n = 1
    align = vb.styles.alignment.Alignment(horizontal='center', vertical='center')
    table.cell(n, 1, value=datetime.date.today().strftime('%Y-%m-%d'))
    table.cell(n, 1).font=normal_font
    table.row_dimensions[n].height = 20.4
    n += 1
    for key, value in data.items():
        table.cell(n, 1, value=key)
        table.cell(n, 1).font = title_font
        table.merge_cells('A%s:B%s' % (n, n))
        table.cell(n, 1).alignment = align
        table.cell(n, 1).border = border
        table.cell(n, 2).border = border
        table.row_dimensions[n].height = 28.2
        n += 1

        table.cell(n, 1, value='预定时间')
        table.cell(n, 1).alignment = align
        table.cell(n, 1).border = border
        table.cell(n, 1).font = subheading_font
        table.cell(n, 2, value='学号')
        table.cell(n, 2).alignment = align
        table.cell(n, 2).border = border
        table.cell(n, 2).font = subheading_font
        table.row_dimensions[n].height = 30
        n += 1

        for id, timetable in value.items():
            for info_key, info_value in timetable.items():
                if info_key == 'start_time':
                    start_time = info_value.strftime('%H:%M')
                if info_key == 'end_time':
                    table.cell(n, 1, value=start_time + '——' + info_value.strftime('%H:%M'))
                    table.cell(n, 1).alignment = align
                    table.cell(n, 1).border = border
                    table.cell(n, 1).font = normal_font
                if info_key == 'users':
                    if info_value:
                        table.merge_cells('A%s:A%s' % (n, n - 1 + len(timetable['users'])))
                    for ticketid,users in info_value.items():
                        for user in users:
                            table.cell(n, 2, value=user)
                            table.cell(n, 2).alignment = align
                            table.cell(n, 2).border = border
                            table.cell(n, 2).font = normal_font
                            table.cell(n, 3, value=ticketid)
                            table.cell(n, 3).alignment = align
                            table.cell(n, 3).border = border
                            table.cell(n, 3).font = normal_font
                            table.row_dimensions[n].height = 20.4
                            n += 1
                if not info_value:
                    table.cell(n, 2).border = border
                    n += 1
        n += 1
    file_name=datetime.date.today().strftime("%Y-%m-%d")+'.xlsx'
    excel.save(f'./static/excel/'+file_name)
    return file_name
